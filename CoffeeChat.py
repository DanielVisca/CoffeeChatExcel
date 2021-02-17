import csv
import random
import json
import openpyxl
from openpyxl import Workbook
from pathlib import Path
import os
import io
import tkinter as tk
from tkinter import messagebox
# os.chdir(sys._MEIPASS)
# os.system('CoffeeChat.xlsx')

class CoffeeChat:
    """
    db format:
    {
        people: [Person Object, ...],
        previous_matches: [Match_object]
    }
    """
    def __init__(self):
        self.excel_path = os.getcwd() + "/Desktop/CoffeeChat/CoffeeChat.xlsx"
        self.window = tk.Tk()
        self.gui()

    def gui(self):
        label1 = tk.Label(text="Hello, Aidan! Coffee Chat time :)")
        label1.pack()

        frame1 = tk.Frame()
        instructions = tk.Label(text="\nInstructions:")
        instruction_list = [
            tk.Label(frame1, text="1) There must be a folder on your desktop labelled 'CoffeeChat'", anchor="w", width=100),
            tk.Label(frame1, text="2) Press 'Generate new partners' for the first time for an empty excel workbook to be created", anchor="w", width=100),
            tk.Label(frame1, text="3) In the workbook there is a sheet called 'enrolled', the first column is for emails, 2nd for first name, 3rd for last name, 4th for position", anchor="w", width=100),
            tk.Label(frame1, text="4) You can update the enrolled list at anytime", anchor="w", width=100),
            tk.Label(frame1, text="5) Select 'Generate new partners' then reload the excel workbook. A new sheet will be made with the matches.", anchor="w", width=100),
            tk.Label(frame1, text="6) step 5 can be repeated until no new matches can be made.", anchor="w", width=100),
            tk.Label(frame1, text="7) To reset the match history, select 'Delete Database' there is no way to restore deleted match history and no way to re-roll a week.", anchor="w", width=100)
        ]
        for item in instruction_list:
            item.pack()
        frame1.pack()   
        button = tk.Button(
            text="Generate new partners!",
            width=25,
            height=2,
            command=self.random_match
        )
        button1 = tk.Button(
            text="Delete Database (clear match history)",
            width=30,
            height=2,
            command=self.delete_db
        )
        
        button.pack()
        button1.pack()
        self.window.mainloop()
        # self.print_matches()
        # print(self.test())
    
    def delete_db(self):
        msg_box = messagebox.askokcancel("askokcancel", "Are you sure you want to delete all match history?")
        if os.path.exists("database.json"):
            if msg_box:
                os.remove("database.json")
                print('database deleted')
            else:
                messagebox.showinfo('Return','Nothing was deleted, you will now return to the application screen')
        else:
            print('No database found')

    def save(self):
        data = {
            'people': [person for person in self.people],
            'previous_matches': [match for match in self.previous_matches]
        }
        with open('database.json','w') as f: 
            json.dump(data, f, indent=4) 
        
        self.write()
    

    def update_people(self):
        """
        Requires more extensive testing!

        If there are new people, create them and add them to the list.
        If people have been removed. remove them from the list
        """
        people = self.get_people_excel()
        # check excel against db list of people. This will find new additions
        for person in people: 
            if person not in self.people: # uncertain if 'in' will work with custom class
                self.people.append(person)
        
        # check sb list against updated excel. This will find removed people?
        for person in self.people:
            if person not in people:
                self.people.remove(person)

    def get_people_excel(self):
        try:
            xlsx_file = Path(self.excel_path)
            self.wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet = self.wb_obj['enrolled']
            people = []
            for row in sheet.iter_rows(max_row=sheet.max_row):
                person = Person(row[0].value, row[1].value, row[2].value ,row[3].value)
                people.append(person)
            return people
        except:
            print("No workbook found, creating one")
            self.wb_obj = Workbook()
            self.wb_obj.create_sheet("enrolled")
            self.wb_obj.save(self.excel_path)
            return []

    def upload(self):
        if os.path.isfile('database.json') and os.access("database.json", os.R_OK):
            print("database exists")
            with open('database.json') as f:
                db = json.load(f)
                try:
                    self.people = list(db['people'])
                    self.previous_matches = list(db['previous_matches'])
                    print('history successfully retrieved from database')
                except:
                    print("Could not read 'People' and 'Previous Matches' from database. Creating empty lists")
                    self.people = []
                    self.previous_matches = []

        else:
            print ("The database does not exit, creating a new one...")
            with io.open(os.path.join('database.json'), 'w') as db_file:
                db_file.write(json.dumps({}))
                self.people = []
                self.previous_matches = []
    
    def match(self, p1, p2):
        """
        Check if p1 and p2 have been matched before. If they have, return None, if they have not, create a new Match obj and
        add it to self.matches (which will be added to db)
        """
        new_match = Match(p1, p2)
        e3, e4 = new_match['emails']
        for match in self.previous_matches:
            e1, e2 = match['emails']
            
            # also set( of size 2 or less)
            # Check if both emails match
            if ((e1 == e3) or (e1 == e4)) and ((e2 == e3) or (e2 == e4)):
                return False
    
        return new_match
    

    def random_match(self):
        """
        Randomly match email addresses
        """
        self.upload() # creates self.people and self.previous_matches
        self.update_people()

        match_list = []
        people_to_match = self.people.copy()
        unmatched = None

        while len(people_to_match) > 1:
            person = random.choice(people_to_match)
            # can I remove a whole dict from a list?
            people_to_match.remove(person)

            random_partner = random.choice(people_to_match)
            match = self.match(person, random_partner)

            seen_everyone = []
            tried = [random_partner]
            while not match: # already matched previously
            # maybe shift this to lower in the while loop
                if len(tried) >= len(people_to_match):
                    seen_everyone.append(person)
                    print(person['email'], " has matched with everyone already and cannot find a partner")
                    break
                random_partner = random.choice(people_to_match)
                if random_partner not in tried:
                    tried.append(random_partner)
                    match = self.match(person, random_partner)
            if match:
                # By this point they have never been matched before
                people_to_match.remove(random_partner)
                good_match = Match(person, random_partner)
                match_list.append(good_match)
                self.previous_matches.append(good_match)
        
        self.match_list = match_list
        self.save()

    def print_matches(self):
        print("\nThis weeks matches:\n")
        for match in self.match_list:
            print(match["emails"])
    
    def test(self):
        possible_matches = 0
        for i in range(1, len(self.people)):
            possible_matches += i
        return len(self.previous_matches) == possible_matches
    
    def write(self):
        ws1 = self.wb_obj.create_sheet("week")
        for match in self.match_list:
            people1 = match['people'][0]
            people2 = match['people'][1]
            row = (people1['email'], people1['first_name'], people1['last_name'], people1['position'], '', people2['email'], people2['first_name'], people2['last_name'], people2['position'])
            ws1.append(row)
        self.wb_obj.save(self.excel_path)
        messagebox.showinfo('Saved','Reload your excel (close and open) to see a new worksheet')
        print("Matches saved to excel!")

def Person(email, first_name, last_name, position):
    return {
        'email': email,
        'first_name': first_name,
        'last_name': last_name,
        'position': position
    }

def Match(p1,p2):
    return {
        "emails": (p1['email'], p2['email']),
        "people": (p1, p2)
    }

if __name__ == '__main__':
    cc = CoffeeChat()
